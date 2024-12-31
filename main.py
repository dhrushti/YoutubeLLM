# from flask import Flask, render_template, request
# import csv
# import os

# app = Flask(__name__)

# # Path to the CSV file
# CSV_FILE = "product_data.csv"

# # Ensure the CSV file exists and has headers
# if not os.path.exists(CSV_FILE):
#     with open(CSV_FILE, mode="w", newline="") as file:
#         writer = csv.writer(file)
#         writer.writerow(["Product", "Problem", "Solution"])  # CSV headers

# @app.route("/", methods=["GET", "POST"])
# def form():
#     if request.method == "POST":
#         # Retrieve form data
#         product = request.form.get("product")
#         problem = request.form.get("problem")
#         solution = request.form.get("solution")

#         # Append data to the CSV file
#         with open(CSV_FILE, mode="a", newline="") as file:
#             writer = csv.writer(file)
#             writer.writerow([product, problem, solution])

#         # Provide user feedback
#         return render_template("success.html")

#     # Render the input form
#     return render_template("form.html")

# if __name__ == "__main__":
#     app.run(debug=True)

from flask import Flask, render_template, request
import os
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Path to the Excel file
EXCEL_FILE = "product_data.xlsx"

# Ensure the Excel file exists and has headers
def create_excel_file():
    if not os.path.exists(EXCEL_FILE):
        # Create a new workbook and add headers
        wb = Workbook()
        ws = wb.active
        ws.append(["Product", "Problem", "Solution"])  # Headers
        wb.save(EXCEL_FILE)

# Function to append data to the Excel file
def append_to_excel(product, problem, solution):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([product, problem, solution])
    wb.save(EXCEL_FILE)

# Ensure the Excel file exists when the app starts
create_excel_file()

@app.route("/", methods=["GET", "POST"])
def form():
    if request.method == "POST":
        # Retrieve form data
        product = request.form.get("product")
        problem = request.form.get("problem")
        solution = request.form.get("solution")

        # Append data to the Excel file
        append_to_excel(product, problem, solution)

        # Provide user feedback
        return render_template("success.html")

    # Render the input form
    return render_template("form.html")

if __name__ == "__main__":
    app.run(debug=True)
